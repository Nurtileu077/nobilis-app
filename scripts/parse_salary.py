#!/usr/bin/env python3
"""
Parse salary Excel files and generate salaryData.js for the Nobilis Academy app.
"""

import openpyxl
import json
import re
import datetime

STAFF_FILE = 'data/salary/ЗП Nobilis Сотрудников (2).xlsx'
SALES_FILE = 'data/salary/ЗП ОП Nobiis (1).xlsx'
OUTPUT_FILE = 'src/data/salaryData.js'

# Sheet name → formatted month label
# The file contains sheets: Июнь, Июль, Август, Сентябрь, Октябрь, Ноябрь, Декабрь (all 2024)
# then Январь, Февраль, Март, Апрель, Май (2025)
# then Июнь 2025, Июль 2025, Август 2025, etc.
MONTH_MAP_2024 = {
    'Июнь':     'Июнь 2024',
    'Июль':     'Июль 2024',
    'Август':   'Август 2024',
    'Сентябрь': 'Сентябрь 2024',
    'Октябрь':  'Октябрь 2024',
    'Ноябрь':   'Ноябрь 2024',
    'Декабрь':  'Декабрь 2024',
    'Январь':   'Январь 2025',
    'Февраль':  'Февраль 2025',
    'Март':     'Март 2025',
    'Апрель':   'Апрель 2025',
    'Май':      'Май 2025',
}

def normalize_sheet_name(shname):
    """Convert sheet name to full month+year label."""
    shname_stripped = shname.strip()
    # Already has year info (e.g. "Июнь 2025", "Декабрь 2025 ")
    if re.search(r'20\d\d', shname_stripped):
        # Normalize spacing
        return re.sub(r'\s+', ' ', shname_stripped).strip()
    # Bare month name -> 2024 or early 2025
    return MONTH_MAP_2024.get(shname_stripped, shname_stripped)


def infer_role(name):
    """Infer employee role from name."""
    if not name:
        return 'other'
    n = name.strip()
    # Directors / founders
    if re.search(r'Ерболат|Аружан.*учредитель|Аружан бывш', n, re.IGNORECASE):
        return 'director'
    if re.search(r'Нұртілеу|Нуртилеу|Тұрсынхан', n, re.IGNORECASE):
        return 'director'
    # Academic director
    if re.search(r'Байғазы|Салтанат', n, re.IGNORECASE):
        return 'academic_director'
    # Curators
    if re.search(r'Расиль|Асхат', n, re.IGNORECASE):
        return 'curator'
    if re.search(r'^Дархан\b', n, re.IGNORECASE):
        return 'curator'
    # Coordinators
    if re.search(r'^Нурым\b', n, re.IGNORECASE):
        return 'coordinator'
    if re.search(r'^Бейбут\b|^Бейбит\b', n, re.IGNORECASE):
        return 'coordinator'
    if re.search(r'Аруна.*ПУ|^Аруна\b', n, re.IGNORECASE):
        return 'coordinator'
    # Sales managers
    if re.search(r'^Мадияр\b', n, re.IGNORECASE):
        return 'sales_manager'
    if re.search(r'^Мадина\b', n, re.IGNORECASE):
        return 'sales_manager'
    if re.search(r'^Анель\b', n, re.IGNORECASE):
        return 'sales_manager'
    # ROP (head of sales)
    if re.search(r'Алмат.*РОП|РОП', n, re.IGNORECASE):
        return 'rop'
    # Call center
    if re.search(r'^Руслан\b', n, re.IGNORECASE):
        return 'callcenter'
    if re.search(r'^Динара\b|Динара колл|^Динара к', n, re.IGNORECASE):
        return 'callcenter'
    if re.search(r'^Камиля\b', n, re.IGNORECASE):
        return 'callcenter'
    if re.search(r'Асель колл\b|Асель к', n, re.IGNORECASE):
        return 'callcenter'
    # Office manager / cleaning
    if re.search(r'Эльвира|уборщица', n, re.IGNORECASE):
        return 'office_manager'
    # Accountant
    if re.search(r'Бухгалтер|Гульжахан', n, re.IGNORECASE):
        return 'accountant'
    # Teachers
    if re.search(r'Алена.*Гугл|Алена Гугл|^Алена\b', n, re.IGNORECASE):
        return 'teacher'
    if re.search(r'Алуа.*Ментор|Алуа.*Айлтс|Алуа.*Ielts|^Алуа\b', n, re.IGNORECASE):
        return 'teacher'
    if re.search(r'^Бекмухамед\b', n, re.IGNORECASE):
        return 'teacher'
    if re.search(r'Асель SAT|Асель Сат', n, re.IGNORECASE):
        return 'teacher'
    # Other roles
    if re.search(r'Маргулан.*Видео|Видеограф|^Султан\b', n, re.IGNORECASE):
        return 'other'
    if re.search(r'СММ|SMM|Абылай|Еркежан|Арайлым.*СММ|Малика.*СММ|Малика\b', n, re.IGNORECASE):
        return 'other'
    if re.search(r'Құндызай|Бахадыр.*Юрист|Юрист', n, re.IGNORECASE):
        return 'other'
    if re.search(r'Амина.*Дизайнер|Аружан.*Дизайнер|Дизайнер', n, re.IGNORECASE):
        return 'other'
    if re.search(r'^Али\b', n, re.IGNORECASE):
        return 'other'
    if re.search(r'^Улан\b', n, re.IGNORECASE):
        return 'other'
    return 'other'


def safe_num(val):
    """Convert value to number, treating None/formulas/strings as 0."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    if isinstance(val, str):
        # Try to parse number from string
        cleaned = val.replace(' ', '').replace(',', '.')
        try:
            return round(float(cleaned), 2)
        except ValueError:
            return 0
    if isinstance(val, datetime.datetime):
        return 0
    return 0


def normalize_name(name):
    """Strip and normalize employee name."""
    if not name:
        return None
    name = str(name).strip()
    # Remove trailing/leading whitespace
    return name if name else None


def parse_staff_file(filepath):
    """Parse staff salary file and return employees dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # employees: {normalized_name: {role, salaries: {month_label: {...}}}}
    employees = {}
    month_labels = []

    for shname in wb.sheetnames:
        month_label = normalize_sheet_name(shname)
        month_labels.append(month_label)

        sheet = wb[shname]

        # Read header row to determine column positions
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Find column indices
        advance_cols = []
        bonus_col = None
        remaining_col = None
        base_col = 1  # always column B (index 1) = salary

        for i, h in enumerate(header_row):
            if h is None:
                continue
            h_lower = str(h).lower()
            if 'остаток' in h_lower:
                remaining_col = i
            elif 'премия' in h_lower:
                bonus_col = i
            elif 'аванс' in h_lower:
                advance_cols.append(i)

        # Read data rows (skip header, stop at summary/empty rows)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name_raw = row[0]
            if name_raw is None:
                continue

            name = normalize_name(name_raw)
            if not name:
                continue

            # Skip summary rows and work-days section headers
            name_lower = name.lower()
            if any(kw in name_lower for kw in ['общ', 'итого', 'total', 'фио', 'имя', 'рабочие']):
                # 'Общ' marks the end of salary data; stop processing this sheet
                if 'общ' in name_lower or 'итого' in name_lower:
                    break
                continue

            # Skip placeholder names
            if name.strip() in ['-', '—', '–']:
                continue

            # Skip empty-data rows (all numeric cols are None/0)
            base_salary = safe_num(row[base_col]) if len(row) > base_col else 0

            # Skip rows that appear to be from the 'work days' sub-table
            # (base salary < 1000 is a day count, not a real salary)
            if base_salary < 1000 and base_salary > 0:
                continue

            # Compute advances total
            advances_total = 0
            advances_list = []
            for ac in advance_cols:
                if ac < len(row):
                    v = safe_num(row[ac])
                    advances_total += v
                    advances_list.append(v)

            bonus = safe_num(row[bonus_col]) if bonus_col is not None and bonus_col < len(row) else 0
            remaining = safe_num(row[remaining_col]) if remaining_col is not None and remaining_col < len(row) else 0

            # Canonicalize name for grouping (strip suffixes like "бывш учредитель", "директор владелец")
            # but keep the display name clean
            canonical = canonicalize_name(name)

            if canonical not in employees:
                employees[canonical] = {
                    'name': canonical,
                    'role': infer_role(canonical),
                    'salaries': {}
                }

            employees[canonical]['salaries'][month_label] = {
                'base': base_salary,
                'advances': round(advances_total, 2),
                'advancesList': [round(a, 2) for a in advances_list if a > 0],
                'bonus': bonus,
                'remaining': remaining,
            }

    return employees, month_labels


def canonicalize_name(name):
    """Remove common suffixes and normalize name for grouping."""
    if not name:
        return name
    # Remove trailing parentheticals like "(уборщица)", "(ПУ)"
    name = re.sub(r'\s*\(.*?\)\s*', ' ', name).strip()
    # Remove suffixes that appear in some months
    suffixes = [
        r'\s+бывш\s+учредитель.*$',
        r'\s+директор\s+владелец.*$',
        r'\s+учредитель.*$',
        r'\s+владелец.*$',
        r'\s+колл\s+центр.*$',
        r'\s+колл\s*$',
        r'\s+к\s+ц.*$',
    ]
    for s in suffixes:
        name = re.sub(s, '', name, flags=re.IGNORECASE).strip()
    return name


# Sales file: blocks of 27 rows each, 3 managers side by side
# Madiyar: cols A-L (0-11)
# Madina:  cols M-X (12-23)
# Anel:    cols Y-AJ (24-35)
# Row offsets within each block:
#   0: Manager header
#   1: Month name
#   2: Column headers
#   3: First data row = totals row (H/T/AF = total sales, I/U/AG = net sales, J/V/AH=3%, K/W/AI=5%, L/X/AJ=7%)
#   4: 'ОБЩ к-лво Продаж' label row (no numeric total deal count stored here)
#   5+: More data rows with individual deal commissions or empty

SALES_MONTH_SEQUENCE = [
    # blocks 1-7 = June-Dec 2024
    'Июнь 2024', 'Июль 2024', 'Август 2024', 'Сентябрь 2024', 'Октябрь 2024', 'Ноябрь 2024', 'Декабрь 2024',
    # blocks 8-12 = Jan-May 2025
    'Январь 2025', 'Февраль 2025', 'Март 2025', 'Апрель 2025', 'Май 2025',
    # blocks 13-19 = June 2025 - Dec 2025
    'Июнь 2025', 'Июль 2025', 'Август 2025', 'Сентябрь 2025', 'Октябрь 2025', 'Ноябрь 2025', 'Декабрь 2025',
    # blocks 20-22 = duplicated / extra sheets in staff file, but in sales: Январь 2026, Февраль 2026, Март 2026
    'Январь 2026', 'Февраль 2026', 'Март 2026',
]


def parse_sales_month_name(raw):
    """Convert raw month name from sales sheet to full label."""
    if raw is None:
        return None
    raw = str(raw).strip()
    # Already has year
    if re.search(r'20\d\d', raw):
        return re.sub(r'\s+', ' ', raw).capitalize()
    # Bare month name – use sequence index to determine year
    return raw


def parse_sales_file(filepath):
    """Parse the sales department file and return managers data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['Лист1']
    all_rows = list(sheet.iter_rows(values_only=True))

    BLOCK_SIZE = 27

    managers = {
        'Мадияр': {'name': 'Мадияр', 'role': 'sales_manager', 'sales': {}},
        'Мадина': {'name': 'Мадина', 'role': 'sales_manager', 'sales': {}},
        'Анель':  {'name': 'Анель',  'role': 'sales_manager', 'sales': {}},
    }

    month_labels = []

    num_blocks = len(all_rows) // BLOCK_SIZE

    for block_idx in range(num_blocks):
        start = block_idx * BLOCK_SIZE

        # Row 2 (offset 1) has the month name
        month_raw = all_rows[start + 1][0]
        if month_raw is None:
            continue

        # Determine the full month label
        # Use the SALES_MONTH_SEQUENCE for the first 23 usable blocks
        if block_idx < len(SALES_MONTH_SEQUENCE):
            month_label = SALES_MONTH_SEQUENCE[block_idx]
        else:
            # Fall back to parsing the raw name
            month_label = str(month_raw).strip()

        if month_label not in month_labels:
            month_labels.append(month_label)

        # Row 4 (offset 3): totals row
        totals_row = all_rows[start + 3]

        # Count deals per manager by counting non-None/non-header entries in their respective
        # contract columns across the data rows (offsets 3 to 26)
        def count_deals(col_idx):
            count = 0
            for offset in range(3, BLOCK_SIZE):
                val = all_rows[start + offset][col_idx]
                if val is None:
                    continue
                if isinstance(val, (int, float)):
                    count += 1
                elif isinstance(val, str):
                    s = val.strip()
                    if s and 'договора' not in s.lower() and 'менеджер' not in s.lower() and 'к-лво' not in s.lower():
                        count += 1
                elif isinstance(val, datetime.datetime):
                    count += 1
            return count

        madiyar_deals = count_deals(0)   # col A
        madina_deals = count_deals(12)   # col M
        anel_deals = count_deals(24)     # col Y

        # Total sales: col H (7), col T (19), col AF (31) from totals row
        madiyar_total_sales = safe_num(totals_row[7])
        madina_total_sales = safe_num(totals_row[19])
        anel_total_sales = safe_num(totals_row[31])

        # Net sales after fee deductions: col I (8), col U (20), col AG (32)
        madiyar_net_sales = safe_num(totals_row[8])
        madina_net_sales = safe_num(totals_row[20])
        anel_net_sales = safe_num(totals_row[32])

        # Commission at 5% without pension deduction: col K (10), col W (22), col AI (34)
        madiyar_comm_5pct = safe_num(totals_row[10])
        madina_comm_5pct = safe_num(totals_row[22])
        anel_comm_5pct = safe_num(totals_row[34])

        # Commission at 3%: col J (9), col V (21), col AH (33)
        madiyar_comm_3pct = safe_num(totals_row[9])
        madina_comm_3pct = safe_num(totals_row[21])
        anel_comm_3pct = safe_num(totals_row[33])

        # Commission at 7%: col L (11), col X (23), col AJ (35)
        madiyar_comm_7pct = safe_num(totals_row[11])
        madina_comm_7pct = safe_num(totals_row[23])
        anel_comm_7pct = safe_num(totals_row[35])

        managers['Мадияр']['sales'][month_label] = {
            'totalSales': madiyar_total_sales,
            'netSales': madiyar_net_sales,
            'deals': madiyar_deals,
            'commission3pct': madiyar_comm_3pct,
            'commission5pct': madiyar_comm_5pct,
            'commission7pct': madiyar_comm_7pct,
        }
        managers['Мадина']['sales'][month_label] = {
            'totalSales': madina_total_sales,
            'netSales': madina_net_sales,
            'deals': madina_deals,
            'commission3pct': madina_comm_3pct,
            'commission5pct': madina_comm_5pct,
            'commission7pct': madina_comm_7pct,
        }
        managers['Анель']['sales'][month_label] = {
            'totalSales': anel_total_sales,
            'netSales': anel_net_sales,
            'deals': anel_deals,
            'commission3pct': anel_comm_3pct,
            'commission5pct': anel_comm_5pct,
            'commission7pct': anel_comm_7pct,
        }

    return managers, month_labels


def to_js_value(val, indent=0):
    """Convert Python value to JS literal string."""
    pad = '  ' * indent
    if isinstance(val, dict):
        if not val:
            return '{}'
        lines = ['{']
        items = list(val.items())
        for i, (k, v) in enumerate(items):
            comma = ',' if i < len(items) - 1 else ''
            lines.append(f"{pad}  '{k}': {to_js_value(v, indent + 1)}{comma}")
        lines.append(pad + '}')
        return '\n'.join(lines)
    elif isinstance(val, list):
        if not val:
            return '[]'
        if all(isinstance(x, (int, float, str)) for x in val):
            inner = ', '.join(to_js_value(x) for x in val)
            return f'[{inner}]'
        lines = ['[']
        for i, x in enumerate(val):
            comma = ',' if i < len(val) - 1 else ''
            lines.append(f"{pad}  {to_js_value(x, indent + 1)}{comma}")
        lines.append(pad + ']')
        return '\n'.join(lines)
    elif isinstance(val, str):
        escaped = val.replace('\\', '\\\\').replace("'", "\\'")
        return f"'{escaped}'"
    elif isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return repr(round(val, 2))
    elif isinstance(val, int):
        return str(val)
    elif val is None:
        return 'null'
    elif isinstance(val, bool):
        return 'true' if val else 'false'
    return repr(val)


def format_employee_js(emp, all_months):
    """Format a single employee as a JS object literal."""
    lines = []
    lines.append('  {')
    lines.append(f"    name: '{emp['name']}',")
    lines.append(f"    role: '{emp['role']}',")
    lines.append('    salaries: {')

    for month in all_months:
        if month in emp['salaries']:
            s = emp['salaries'][month]
            advances_list_str = '[' + ', '.join(str(int(a) if a == int(a) else a) for a in s['advancesList']) + ']'
            base = int(s['base']) if s['base'] == int(s['base']) else s['base']
            advances = int(s['advances']) if s['advances'] == int(s['advances']) else s['advances']
            bonus = int(s['bonus']) if s['bonus'] == int(s['bonus']) else s['bonus']
            remaining = int(s['remaining']) if s['remaining'] == int(s['remaining']) else s['remaining']
            lines.append(f"      '{month}': {{ base: {base}, advances: {advances}, advancesList: {advances_list_str}, bonus: {bonus}, remaining: {remaining} }},")

    lines.append('    },')
    lines.append('  }')
    return '\n'.join(lines)


def format_manager_js(mgr, all_months):
    """Format a single sales manager as a JS object literal."""
    lines = []
    lines.append('  {')
    lines.append(f"    name: '{mgr['name']}',")
    lines.append(f"    role: '{mgr['role']}',")
    lines.append('    sales: {')

    for month in all_months:
        if month in mgr['sales']:
            s = mgr['sales'][month]
            def fmt(v):
                if isinstance(v, float) and v == int(v):
                    return str(int(v))
                return repr(round(v, 2)) if isinstance(v, float) else str(v)
            lines.append(
                f"      '{month}': {{ "
                f"totalSales: {fmt(s['totalSales'])}, "
                f"netSales: {fmt(s['netSales'])}, "
                f"deals: {s['deals']}, "
                f"commission3pct: {fmt(s['commission3pct'])}, "
                f"commission5pct: {fmt(s['commission5pct'])}, "
                f"commission7pct: {fmt(s['commission7pct'])} "
                f"}},"
            )

    lines.append('    },')
    lines.append('  }')
    return '\n'.join(lines)


def main():
    print("Parsing staff salary file...")
    employees, staff_month_labels = parse_staff_file(STAFF_FILE)

    print(f"Found {len(employees)} employees across {len(staff_month_labels)} months")
    print("Months:", staff_month_labels)

    print("\nParsing sales file...")
    managers, sales_month_labels = parse_sales_file(SALES_FILE)

    print(f"Found {len(managers)} managers across {len(sales_month_labels)} months")

    # Build the JS output
    lines = []
    lines.append('// Auto-generated from Excel salary files')
    lines.append('// Staff file: ЗП Nobilis Сотрудников (2).xlsx')
    lines.append('// Sales file: ЗП ОП Nobiis (1).xlsx')
    lines.append('')

    # SALARY_DATA
    months_str = '[' + ', '.join(f"'{m}'" for m in staff_month_labels) + ']'
    lines.append('export const SALARY_DATA = {')
    lines.append(f'  months: {months_str},')
    lines.append('  employees: [')

    emp_list = sorted(employees.values(), key=lambda e: e['name'])
    for i, emp in enumerate(emp_list):
        emp_js = format_employee_js(emp, staff_month_labels)
        if i < len(emp_list) - 1:
            emp_js += ','
        lines.append(emp_js)

    lines.append('  ],')
    lines.append('};')
    lines.append('')

    # SALES_DATA
    sales_months_str = '[' + ', '.join(f"'{m}'" for m in sales_month_labels) + ']'
    lines.append('export const SALES_DATA = {')
    lines.append(f'  months: {sales_months_str},')
    lines.append('  managers: [')

    mgr_list = [managers['Мадияр'], managers['Мадина'], managers['Анель']]
    for i, mgr in enumerate(mgr_list):
        mgr_js = format_manager_js(mgr, sales_month_labels)
        if i < len(mgr_list) - 1:
            mgr_js += ','
        lines.append(mgr_js)

    lines.append('  ],')
    lines.append('};')
    lines.append('')

    output = '\n'.join(lines)

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(output)

    print(f"\nWritten to {OUTPUT_FILE}")
    print(f"File size: {len(output)} bytes")


if __name__ == '__main__':
    main()
